from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
# from tkinter import messagebox


root = Tk()
root.title('ERIC PY PROGRAM')
root.geometry("400x400")
# messagebox.showinfo("Showinfo", "Information")

# select function
def select(e):
    my_label.config(text = my_listbox.get(ANCHOR))

# my_list = ["one","two","three"]

my_listbox = Listbox(root, width=50)
my_listbox.pack(pady=20)

# for item in my_list:
#     my_listbox.insert(END, item)

# creation de  l instance workbook
wb = load_workbook("name_color.xlsx")

# instance du worksheet, set active worksheet
ws = wb.active

# grab a column of data
col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_listbox.insert(END, item.value)

my_label = Label(root, text='selectionner un nom', font=("Helvetica",18))
my_label.pack(pady=20)            

# creation listbox binding
my_listbox.bind("<ButtonRelease-1>", select)  

root.mainloop()