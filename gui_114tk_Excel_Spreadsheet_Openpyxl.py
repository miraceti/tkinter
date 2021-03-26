from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = Tk()
root.title('ERIC PY_ OPENPYXL')
root.geometry("500x600")

#creation de l'instance workbook
wb = Workbook()

#ouverture d'un XLS existant
wb = load_workbook('pizza.xlsx')

#creation de la feuille active
ws = wb.active

#creation de vartiable pour la colonne A
column_a = ws['A']
column_b = ws['B']

# print(column_a)

def get_a():
    list =''
    for cell in column_a:
        print(cell.value)
        list =f'{list + str(cell.value)}\n'

    label_a.config(text=list)

def get_b():
    list =''
    for cell in column_b:
        print(cell.value)
        list =f'{list + str(cell.value)}\n'

    label_b.config(text=list)



ba = Button(root, text="Get Column A", command=get_a)
ba.pack(pady=20)

label_a = Label(root, text="")
label_a.pack(pady=20)

bb = Button(root, text="Get Column B", command=get_b)
bb.pack(pady=20)

label_b = Label(root, text="")
label_b.pack(pady=20)


#ajout de fred en A9 et B9
ws['A9'] = "Fred"
ws['B9'] = "Champignons"

#save nouveau fichier
wb.save('pizza2.xlsx')






root.mainloop()